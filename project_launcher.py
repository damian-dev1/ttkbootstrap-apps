import subprocess
import xlwings as xw
from pathlib import Path
import datetime
import subprocess
from pathlib import Path
from tkinter import StringVar, Text, END, messagebox, filedialog
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from pathlib import Path
import os
from tkinter import messagebox
import shutil
import time
from openpyxl import load_workbook

SCRIPT_PATH = Path(r"C:/Users/damian/projects/newproject/projectgen.py")
EXCEL_PATH = Path(r"C:/Users/damian/projects/tests/simulate_d/project_dashboard.xlsm")
TEMPLATE_PATH = Path("assets/project_dashboard_template.xlsm")
TEMPLATE_CHECK = True

def load_dropdown_values():
    app = xw.App(visible=False)
    wb = app.books.open(str(EXCEL_PATH))
    ws = wb.sheets["DataValidation"]
    values = {}
    headers = ["Type", "Language", "Status"]
    for idx, header in enumerate(headers, start=1):
        col_range = ws.range((2, idx), ws.range((2, idx)).end("down"))
        col_values = [str(c.value).strip() for c in col_range if c.value]
        values[header] = col_values
    app.quit()
    return values

def log_to_excel(name, type_, lang, status, tags, dev, output):
    try:
        # Attach to the already open Excel workbook
        wb = xw.books["project_dashboard.xlsm"]
    except Exception:
        # If not open, open it in background mode
        app = xw.App(visible=False)
        wb = app.books.open(str(EXCEL_PATH))
    ws = wb.sheets["Dashboard"]
    last_row = ws.range("A" + str(ws.cells.last_cell.row)).end("up").row + 1
    ws.range(f"A{last_row}").value = [
        datetime.datetime.now(), name, type_, lang, status, tags, dev, output
    ]
    wb.save()
    if "app" in locals():
        app.quit()

def generate_tags(name, type_, lang):
    parts = set(name.lower().split("-") + [type_.lower(), lang.lower()])
    return ",".join(sorted(parts))

def run_projectgen(name, type_, lang, status, dev, tags):
    cmd = [
        "python", str(SCRIPT_PATH),
        name,
        "--type", type_,
        "--lang", lang,
        "--status", status,
        "--tags", tags,
        "--dev", dev
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)
    return result

def archive_dashboard_rows(excel_path: Path):
    app = xw.App(visible=False)
    wb = app.books.open(str(excel_path))
    try:
        dashboard = wb.sheets["Dashboard"]
        archive = None
        if "Archive" in [s.name for s in wb.sheets]:
            archive = wb.sheets["Archive"]
        else:
            archive = wb.sheets.add("Archive")
        data = dashboard.range("A2").expand("table").value
        if not data:
            print("[INFO] Dashboard is empty, nothing to archive.")
            return
        if not isinstance(data[0], list):
            data = [data]
        archive_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        archived_rows = [[archive_date] + row for row in data]
        last_row = archive.range("A" + str(archive.cells.last_cell.row)).end("up").row
        if last_row == 1 and not archive.range("A1").value:
            # Add headers if first time
            archive.range("A1").value = [
                "Date Archived", "Created", "Project Name", "Type", "Language",
                "Status", "Tags", "Folder Path", "Notes"
            ]
            last_row = 1
        archive.range(f"A{last_row + 1}").value = archived_rows
        wb.save()
        print(f"[INFO] Archived {len(archived_rows)} rows to Archive sheet.")
    except Exception as e:
        print(f"[ERROR] Failed to archive Dashboard: {e}")
    finally:
        wb.close()
        app.quit()

def create_dashboard_workbook(excel_path: Path) -> bool:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Template not found at {TEMPLATE_PATH}")
    if excel_path.exists():
        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, keep_vba=True)
            if "Dashboard" in wb.sheetnames:
                dashboard = wb["Dashboard"]
                archive_name = f"Dashboard_Archive_{time.strftime('%Y%m%d_%H%M%S')}"
                wb.copy_worksheet(dashboard).title = archive_name
                wb.save(excel_path)
                print(f"[INFO] Archived Dashboard to '{archive_name}'")
        except Exception as e:
            print(f"[WARNING] Failed to archive dashboard: {e}")
    shutil.copy(TEMPLATE_PATH, excel_path)
    return True

def merge_archived_logs(excel_path: Path):
    import xlwings as xw
    app = xw.App(visible=False)
    wb = app.books.open(str(excel_path))
    try:
        archive_sheet = None
        for sht in wb.sheets:
            if sht.name.startswith("Dashboard_Archive_"):
                archive_sheet = sht
                break
        if not archive_sheet:
            return
        data = archive_sheet.range("A2").expand("table").value
        if not data:
            return
        dashboard = wb.sheets["Dashboard"]
        last_row = dashboard.range("A" + str(dashboard.cells.last_cell.row)).end("up").row + 1
        dashboard.range(f"A{last_row}").value = data
        wb.save()
        print(f"[INFO] Merged {len(data)} rows from archive.")

    except Exception as e:
        print(f"[ERROR] Failed to merge archive logs: {e}")

    finally:
        wb.close()
        app.quit()

if not EXCEL_PATH.exists():
    print("[INFO] Dashboard workbook not found. Creating from template...")
    created = create_dashboard_workbook(EXCEL_PATH)
    if not created:
        from tkinter import messagebox
        messagebox.showerror("Fatal Error", "Failed to create Excel dashboard workbook from template.\nExiting.")
        raise SystemExit(1)

app = ttk.Window(title="New Project Generator", themename="superhero", size=(700, 400))
app.place_window_center()
name_var = StringVar()
type_var = StringVar(value="Backend")
lang_var = StringVar(value="Python")
status_var = StringVar(value="Planning")
dev_var = StringVar()
dropdowns = load_dropdown_values()
frm_main = ttk.Frame(app, padding=15)
frm_main.pack(fill=BOTH, expand=YES)
frm_left = ttk.Frame(frm_main)
frm_left.grid(row=0, column=0, sticky="nsew", padx=10)
frm_right = ttk.Frame(frm_main)
frm_right.grid(row=0, column=1, sticky="nsew", padx=10)
frm_main.columnconfigure(0, weight=3)
frm_main.columnconfigure(1, weight=1)
def browse_folder():
    path = filedialog.askdirectory()
    if path:
        dev_var.set(path)

def show_help():
    help_text = (
        "[HELP] Keyboard Shortcuts:\n"
        "  ↵  Enter         → Create Project\n"
        "  F1              → Show Help\n"
        "  F2              → Browse Dev Path\n"
        "  F3              → Generate Excel Workbook\n"
        "  F4              → About\n"
        "  Ctrl+Q          → Quit\n\n"
        "Fields:\n"
        "- Project Name (kebab-case)\n"
        "- Type, Language, Status from dropdowns\n"
        "- Dev Path (where code will go)"
    )
    log(help_text, "info")
    messagebox.showinfo("Help", help_text)

def run_script_on_enter(event):
    run_script()

app.bind("<Return>", run_script_on_enter)
ttk.Label(frm_left, text="Project Name").grid(row=0, column=0, sticky=W, pady=5)
ttk.Entry(frm_left, textvariable=name_var).grid(row=0, column=1, sticky=EW)
ttk.Label(frm_left, text="Project Type").grid(row=1, column=0, sticky=W, pady=5)
ttk.Combobox(frm_left, textvariable=type_var, values=dropdowns["Type"]).grid(row=1, column=1, sticky=EW)
ttk.Label(frm_left, text="Language").grid(row=2, column=0, sticky=W, pady=5)
ttk.Combobox(frm_left, textvariable=lang_var, values=dropdowns["Language"]).grid(row=2, column=1, sticky=EW)
ttk.Label(frm_left, text="Status").grid(row=3, column=0, sticky=W, pady=5)
ttk.Combobox(frm_left, textvariable=status_var, values=dropdowns["Status"]).grid(row=3, column=1, sticky=EW)
ttk.Label(frm_left, text="Dev Path").grid(row=4, column=0, sticky=W, pady=5)
ttk.Entry(frm_left, textvariable=dev_var).grid(row=4, column=1, sticky=EW)
frm_left.columnconfigure(1, weight=1)
frm_log = ttk.Frame(app, padding=(15, 0, 15, 15))
frm_log.pack(fill=BOTH, expand=YES)
txt_log = Text(frm_log, height=6, bg="#1E1E1E", fg="#DCDCDC", insertbackground="#DCDCDC",
               font=("Consolas", 9), wrap="word")
txt_log.pack(fill=BOTH, expand=YES)
scrollbar = ttk.Scrollbar(frm_log, orient="vertical", command=txt_log.yview)
txt_log.config(yscrollcommand=scrollbar.set)
scrollbar.pack(side="right", fill="y")
footer = ttk.Label(app, text="↵ Create | F1 Help | F2 Browse | F3 Excel | F4 About | Ctrl+Q Exit", anchor="center")
footer.pack(side="bottom", fill=X, padx=10, pady=3)
txt_log.tag_configure("timestamp", foreground="#808080")
txt_log.tag_configure("info", foreground="#00FF00")
txt_log.tag_configure("error", foreground="#FF5555")
txt_log.tag_configure("warning", foreground="#FFFF00")
txt_log.tag_configure("success", foreground="#00FF00")

def show_about():
    messagebox.showinfo("About", "📁 Project Generator v1.0\n"
        "Created by Damian D.\n\n"
        "Automates project bootstrapping, logging, and tracking.\n"
        "Enforces naming conventions and provides Excel integration.")

def log(message, tag="info"):
    ts = datetime.datetime.now().strftime("%H:%M:%S")
    txt_log.insert(END, f"[{ts}] {message}\n", tag)
    txt_log.see(END)

def run_script():
    name = name_var.get().strip()
    dev = dev_var.get().strip()

    if not name or not dev:
        log("[ERROR] Project Name and Dev Path are required.", "error")
        return
    type_ = type_var.get()
    lang = lang_var.get()
    status = status_var.get()
    tags = generate_tags(name, type_, lang)
    result = run_projectgen(name, type_, lang, status, dev, tags)
    output = result.stdout.strip() if result.returncode == 0 else result.stderr.strip()
    log_to_excel(name, type_, lang, status, tags, dev, output)
    tag = "success" if result.returncode == 0 else "error"
    log(output, tag)

def reset_excel():
    from tkinter import messagebox
    import xlwings as xw
    from datetime import datetime
    confirm = messagebox.askyesno("Reset Dashboard", "Move all rows to Archive and clear the Dashboard?")
    if not confirm:
        log("[INFO] Reset cancelled.", "warning")
        return
    try:
        wb = None
        excel_filename = EXCEL_PATH.name
        for book in xw.apps.active.books:
            if book.name == excel_filename:
                wb = book
                break
        if not wb:
            app = xw.App(visible=False)
            wb = app.books.open(str(EXCEL_PATH))
        else:
            app = None
        dashboard = wb.sheets["Dashboard"]
        archive = wb.sheets["Archive"] if "Archive" in [s.name for s in wb.sheets] else wb.sheets.add("Archive")
        data = dashboard.range("A2").expand("table").value
        if not data:
            log("[INFO] Dashboard is empty. Nothing to archive.", "warning")
            if app:
                wb.close()
                app.quit()
            return
        if not isinstance(data[0], list):
            data = [data]

        archive_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        archived_rows = [[archive_date] + row for row in data]
        if archive.range("A1").value is None:
            archive.range("A1").value = [
                "Date Archived", "Created", "Project Name", "Type", "Language",
                "Status", "Tags", "Folder Path", "Notes"
            ]
        last_row = archive.range("A" + str(archive.cells.last_cell.row)).end("up").row
        archive.range(f"A{last_row + 1}").value = archived_rows
        dashboard.range("A2:Z1000").clear_contents()
        wb.save()
        if app:
            wb.close()
            app.quit()
        log(f"[INFO] Archived {len(archived_rows)} row(s) and cleared Dashboard.", "success")
    except Exception as e:
        log(f"[ERROR] Failed during reset: {e}", "error")
ttk.Button(frm_right, text="Browse", command=browse_folder, bootstyle=PRIMARY).pack(pady=5, fill=X)
ttk.Button(frm_right, text="Create Project", command=run_script, bootstyle=SUCCESS).pack(pady=5, fill=X)
ttk.Button(frm_right, text="Help", command=show_help, bootstyle=SECONDARY).pack(pady=5, fill=X)
ttk.Button(frm_right, text="Reset Dashboard", command=reset_excel, bootstyle=DANGER).pack(pady=5, fill=X)
    
app.menubar = ttk.Menu(app)
app.menubar.add_command(label="About (F4)", command=show_about)
app.menubar.add_command(label="Exit (Ctrl+Q)", command=app.destroy)
app.config(menu=app.menubar)
app.bind("<Return>", lambda e: run_script())
app.bind("<F1>", lambda e: show_help())
app.bind("<F2>", lambda e: browse_folder())
app.bind("<F3>", lambda e: reset_excel())
app.bind("<F4>", lambda e: show_about())
app.bind("<Control-q>", lambda e: app.destroy())
log("Welcome to Project Generator!", "info")
log("Fill in the fields and click 'Create Project'.", "info")
log("Use 'Generate Excel Sheet' to create a new dashboard.", "info")

app.mainloop()
